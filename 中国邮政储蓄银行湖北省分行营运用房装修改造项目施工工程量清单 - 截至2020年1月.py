from openpyxl import load_workbook
# import numpy as np
# from collections import Counter
# from openpyxl.styles.colors import Color
# from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter , column_index_from_string
# import re
# import random
# from faker import Factory

wb_all = load_workbook('中国邮政储蓄银行湖北省分行营运用房装修改造项目施工工程量清单 - 截至2020年1月.xlsx',data_only=True)

def every_col(floor_num, excel_name):
    sheet_towrite_one_floor = wb_all[floor_num]
    # wb_sub_one_floor = load_workbook(excel_name + '.xlsx', data_only=True)
    wb_sub_one_floor = load_workbook(excel_name + '.xlsx', data_only=True)
    sheet_toread_one_floor = wb_sub_one_floor['表-08 分部分项工程和单价措施项目清单与计价表']

    num_max_toread = sheet_toread_one_floor.max_row
    num_max_towrite = sheet_towrite_one_floor.max_row
    for i in range(1,num_max_toread):
        data_filter = sheet_toread_one_floor['G' + str(i)].value
        if (data_filter != None) and (data_filter != "工程量"):
            for j in range(1, num_max_towrite + 1):
                if sheet_towrite_one_floor['B' + str(j)].value == sheet_toread_one_floor['B' + str(i)].value:
                    sheet_towrite_one_floor['I' + str(j)].value = float(sheet_toread_one_floor['J' + str(i)].value) / float(sheet_towrite_one_floor['K' + str(j)].value)


sheetnames_towrite = wb_all.sheetnames
# for i in range(1,23):
# for i in range(24,29):
#     floor_num = sheetnames_towrite[i]
    # excel_name = sheetnames_towrite[i] + '工程'
    # print(floor_num)
    # every_col(floor_num, floor_num)

# for i in range(1,23):
for i in range(28,29):
    floor_num = sheetnames_towrite[i]
    # excel_name = sheetnames_towrite[i] + '工程'
    print(floor_num)
    every_col(floor_num, floor_num)

wb_all.save(floor_num + 'sample.xlsx')
