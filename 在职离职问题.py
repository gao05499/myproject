from openpyxl import load_workbook
from datetime import datetime
import time
# import numpy as np
# from collections import Counter
# from openpyxl.styles.colors import Color
# from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter , column_index_from_string
# import re
# import random
# from faker import Factory

wb = load_workbook('在职离职问题.xlsx',data_only = True)
sheet_data = wb['员工信息表']
sheet_work = wb['在职']
sheet_leave = wb['离职']

work_list = []
leave_list = []
for i in range(3,7):
    if sheet_data['H' + str(i)].value != None:
        leave_list.append(i)
    else:
        work_list.append(i)

row_work_num = 3
for i in work_list:
    sheet_work['A' + str(row_work_num)].value = sheet_data['A' + str(i)].value
    sheet_work['B' + str(row_work_num)].value = sheet_data['B' + str(i)].value
    sheet_work['C' + str(row_work_num)].value = sheet_data['C' + str(i)].value
    sheet_work['D' + str(row_work_num)].value = sheet_data['D' + str(i)].value
    sheet_work['E' + str(row_work_num)].value = sheet_data['E' + str(i)].value
    sheet_work['F' + str(row_work_num)].value = sheet_data['F' + str(i)].value
    row_work_num += 1

row_leave_num = 3
for i in leave_list:
    sheet_leave['A' + str(row_leave_num)].value = sheet_data['A' + str(i)].value
    sheet_leave['B' + str(row_leave_num)].value = sheet_data['B' + str(i)].value
    sheet_leave['C' + str(row_leave_num)].value = sheet_data['C' + str(i)].value
    sheet_leave['D' + str(row_leave_num)].value = sheet_data['D' + str(i)].value
    sheet_leave['E' + str(row_leave_num)].value = sheet_data['E' + str(i)].value
    sheet_leave['F' + str(row_leave_num)].value = sheet_data['F' + str(i)].value
    sheet_leave['G' + str(row_leave_num)].value = sheet_data['G' + str(i)].value
    sheet_leave['H' + str(row_leave_num)].value = sheet_data['H' + str(i)].value
    sheet_leave['I' + str(row_leave_num)].value = sheet_data['I' + str(i)].value
    row_leave_num += 1

print(work_list)
print(leave_list)

wb.save('sample.xlsx')
