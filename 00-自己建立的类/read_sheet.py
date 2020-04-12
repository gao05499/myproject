# 读取excel文档中所有表单的名称，并输出为列表形式

from openpyxl import load_workbook
def read_sheet(name):
    list_sheet_name = []
    wb = load_workbook(name, data_only=True)

    list_sheet_name = wb.sheetnames  # see all sheet names
    print(list_sheet_name)
    return list_sheet_name
