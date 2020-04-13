# 把一个二维列表转换成一维列表
# 输入有5个参数，输出3个参数
# one_all_data_list_3（带有二维嵌套列表的数据）  [[[]]]
# one_all_data_list_2（带有一维嵌套列表的数据）  [[]]
# one_all_data （不带有列表的数据）  []

from openpyxl import load_workbook
from position import position
from openpyxl.utils import get_column_letter , column_index_from_string
from get_col_alphabet import get_col_alphabet

def from_2D_to_1D(wb_name, sheet_read_name, domain_read, fixed_num, sheet_write):
    wb_name = wb_name
    sheet_name = sheet_read_name
    domain_read = domain_read
    fixed_num = fixed_num
    drift_num = fixed_num + 1
    sheet_write = sheet_write

    wb = load_workbook(wb_name,data_only = True)
    sheet_data = wb[sheet_read_name]
    wb.create_sheet(title=sheet_write,index=0)
    sheet_data_write = wb[sheet_write]
    max_col_num = sheet_data.max_column

    all_data_list, all_data, all_data_list_position, all_data_position, all_data_position_str = position(wb_name,sheet_name,domain_read)

    # 创建表头
    header_col_alphabet = get_col_alphabet(1,fixed_num + 3)
    for i in range(fixed_num):
        col_write = header_col_alphabet[i]
        sheet_data_write[col_write + str(1)].value = all_data_list[0][i]
    sheet_data_write[get_column_letter(fixed_num + 1) + str(1)].value = '需要修改的_1'
    sheet_data_write[get_column_letter(fixed_num + 2) + str(1)].value = '需要修改的_2'

    # 输出结果参数
    one_all_data_list_3 = []
    one_all_data_list_2 = []
    one_all_data = []
    sub_sub_list = []
    sub_list = []

    # 创建数列
    End_col_alphabet = get_col_alphabet(fixed_num + 1,max_col_num + 1)
    row_write = 2
    for i in all_data_list[1:]:
        for j in range(max_col_num - fixed_num):
            for idx_1,k in enumerate(header_col_alphabet[:fixed_num]):
                sheet_data_write[k + str(row_write)].value = i[idx_1]
                sub_sub_list.append(i[idx_1])
                one_all_data.append(i[idx_1])

            sheet_data_write[header_col_alphabet[fixed_num] + str(row_write)].value = all_data_list[0][fixed_num + j]
            sheet_data_write[header_col_alphabet[fixed_num + 1] + str(row_write)].value = i[fixed_num + j]
            sub_sub_list.append(all_data_list[0][fixed_num + j])
            one_all_data.append(all_data_list[0][fixed_num + j])
            sub_sub_list.append(i[fixed_num + j])
            one_all_data.append(i[fixed_num + j])

            one_all_data_list_2.append(sub_sub_list)
            sub_list.append(sub_sub_list)
            sub_sub_list = []

            row_write += 1
        one_all_data_list_3.append(sub_list)
        sub_list = []
    return one_all_data_list_3,one_all_data_list_2,one_all_data

