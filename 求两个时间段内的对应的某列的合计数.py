from openpyxl import load_workbook
from datetime import datetime
# import numpy as np
# from collections import Counter
# from openpyxl.styles.colors import Color
# from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter , column_index_from_string
# import re
# import random
# from faker import Factory

wb_all = load_workbook('求两个时间段内的对应的某列的合计数.xlsx',data_only=True)
sheet_towrite = wb_all['Sheet1']

for j in range(7,12):
    domain_data = sheet_towrite['G' + str(j)].value.split('至')
    data_start = domain_data[0]
    data_end = domain_data[1]
    start_1 = datetime.strptime(data_start, '%Y年%m月%d日')
    end_1 = datetime.strptime(data_end, '%Y年%m月%d日')
    print(data_start + '到' + data_end)

    row_num = sheet_towrite.max_row
    add_cash = 0
    reduce_cash = 0
    for i in range(2,row_num + 1):
        data_read = sheet_towrite['A' + str(i)].value
        if (data_read > start_1) and (data_read <= end_1):
            if (sheet_towrite['B' + str(i)].value != None):
                add_cash += sheet_towrite['B' + str(i)].value
            if (sheet_towrite['C' + str(i)].value != None):
                reduce_cash += sheet_towrite['C' + str(i)].value
                # print(sheet_towrite['B' + str(i)].value)
    
    sheet_towrite['H' + str(j)].value = add_cash
    sheet_towrite['I' + str(j)].value = reduce_cash

wb_all.save('sample.xlsx')
