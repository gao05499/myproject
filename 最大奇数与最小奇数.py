from openpyxl import load_workbook
import numpy as np
from collections import Counter
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter , column_index_from_string
import re
import random
from faker import Factory

wb = load_workbook('最大奇数与最小奇数.xlsx',data_only=True)
sheet_data = wb['Sheet1']
def every_line(start_row, end_row, out_col):
    for i in range(start_row, end_row +1):
        domain_1 = 'C' + str(i) + ':' + 'H' + str(i)
        out_al = out_col + str(i)
        list_col = list(sheet_data[domain_1])
        list_odd = []
        for i in list_col[0]:
            value_1 = i.value
            if value_1 % 2 != 0:
                list_odd.append(value_1)

        list_odd.sort()
        print(list_odd)
        if len(list_odd) >= 2:
            value_result = (list_odd[0] + list_odd[-1]) / 2
        elif len(list_odd) == 1:
            value_result = list_odd[0]
        else:
            value_result = 0
        # print(value_result)


        sheet_data[out_al].value = value_result
    return


every_line(2, 2520, 'K')

wb.save('sample.xlsx')
