from openpyxl import load_workbook
from position import position
from from_2D_to_1D import from_2D_to_1D
from get_col_alphabet import get_col_alphabet

xlsx_name = '连续天数计算次数.xlsx'
sheet_name = 'Sheet1'
domain_read = 'E2:AI3'
fixed_num = 4
sheet_write = '新的列表'

wb = load_workbook(xlsx_name,data_only = True)
sheet_data = wb[sheet_name]
wb.create_sheet(title=sheet_write, index=0)
sheet_data_write = wb[sheet_write]

all_data_list, all_data, all_data_list_position, all_data_position, all_data_position_str = position(xlsx_name,sheet_name,domain_read)

row_num = 5
for i in all_data_list:
    list_1 = []
    num = 0
    for j in i:
        if j == 0:
            list_1.append(num)
            num = 0
        else:
            num += 1
    sheet_data['B' + str(row_num)].value = max(list_1) - 2
    row_num += 1
    print(max(list_1) - 2)

wb.save('sample.xlsx')
