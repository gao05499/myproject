from openpyxl import load_workbook
from openpyxl.styles import  PatternFill
# from datetime import datetime
# import time
# import numpy as np
# from collections import Counter
# from openpyxl.styles.colors import Color
# from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter , column_index_from_string
# import re
# import random
# from faker import Factory
wb = load_workbook('内存数组三之百姓菜篮子.xlsx',data_only = True)
sheet_data = wb['百姓菜篮子']


wb.save('sample.xlsx')
