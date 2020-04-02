from openpyxl import load_workbook
from openpyxl.styles import  PatternFill
# from datetime import datetime
# import time
# import numpy as np
# from collections import Counter
# from openpyxl.styles.colors import Color
# from openpyxl.styles import PatternFill
# from openpyxl.utils import get_column_letter , column_index_from_string
# import re
# import random
# from faker import Factory
wb = load_workbook('内存数组三之百姓菜篮子.xlsx',data_only = True)
sheet_data = wb['百姓菜篮子']


data_id_1 = list(sheet_data['H2:M8'])
all_data = []
sub_data = []
for i in data_id_1:
    for j in i:
        sub_data.append(j.value)
    all_data.append(sub_data)
    sub_data = []

sub_food = []
all_food = []
dic_food = {}
for i in range(2,23):
    food_data = sheet_data['A' + str(i)].value
    for idx,list_1 in enumerate(all_data):
        if food_data in list_1:
            dic_food.setdefault(idx, []).append(food_data)

for key, value in dic_food.items():
    rep_commodity = sheet_data['H' + str(key + 2)].value
    value.remove(rep_commodity)
    value.insert(0, rep_commodity)

fill = PatternFill("solid", fgColor="EEEE00")
row_write = 2
for value in dic_food.values():
    for idx_1 ,food_1 in enumerate(value):
        for i in range(2,23):
            food_eve = sheet_data['A' + str(i)].value
            if food_1 == food_eve:
                if idx_1 == 0:
                    sheet_data['D' + str(row_write)].fill = fill
                    sheet_data['E' + str(row_write)].fill = fill
                    sheet_data['D' + str(row_write)].value = food_eve
                    sheet_data['E' + str(row_write)].value = sheet_data['B' + str(i)].value
                    row_write += 1
                else:
                    sheet_data['D' + str(row_write)].value = food_eve
                    sheet_data['E' + str(row_write)].value = sheet_data['B' + str(i)].value
                    row_write += 1

wb.save('sample.xlsx')
