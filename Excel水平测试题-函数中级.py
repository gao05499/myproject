from openpyxl import load_workbook
import numpy as np
from collections import Counter
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter , column_index_from_string
import re
import random
from faker import Factory

wb = load_workbook('Excel水平测试题.xlsx')
sheet_data1 = wb['4-函数中级']
def get_row_list(start_num, end_start):
    list_1 = list(range(start_num , end_start + 1))
    list_name = []
    for i in list_1:
        list_name.append(sheet_data1['A' + str(i)].value)
    return list_name,list_1

def get_col_list(row_num, start_num, end_start):
    list_1 = list(range(start_num , end_start + 1))
    list_al = []
    list_name = []
    for i in list_1:
        col_al = get_column_letter(i)
        list_al.append(col_al)
        list_name.append(sheet_data1[col_al + str(row_num)].value)
    return list_name,list_al

list_row, list_1 = get_row_list(5,7)
list_col, list_2 = get_col_list(4,2,8)

def search_matrix_list(list_1, list_2):
    search_list_row = []
    search_list_1 = []
    for j in list_1:
        for k in list_2:
            search_list_row.append(sheet_data1[k + str(j)].value)
        search_list_1.append(search_list_row)
        search_list_row = []
    return search_list_1

search_list = search_matrix_list(list_1,list_2)

def index_out(start_num, end_start, search_list, corner_start, corner_end):
    list_3 = list(range(start_num, end_start + 1))

    for i in list_3:
        num_1 = sheet_data1['A' + str(i)].value
        print(num_1)
        search_list = np.array(search_list)
        search_list = search_list - num_1
        search_list = abs(search_list)
        min_num = np.amin(search_list)
        result_index = np.argwhere(search_list == min_num)
        result_index_0 = result_index[0]
        result_index_0 = np.array(result_index_0)
        search_list = result_index_0 + [corner_start,corner_end]
    return search_list

# print(pay_out(10,10,search_list,4,1))
num = index_out(10,10,search_list,5,2)
print(num)

def month_name_out(rd_li_nu,al_ix,num):
    sheet_data1['B' + str(rd_li_nu)].value = sheet_data1[list_2[num[1] - 2] + str(al_ix)].value
    sheet_data1['C' + str(rd_li_nu)].value = sheet_data1['A' + str(list_1[num[0] - 5])].value
    return
month_name_out(10,4,num)

wb.save('sample.xlsx')
