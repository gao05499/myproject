from openpyxl import load_workbook
from collections import Counter
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter , column_index_from_string
import re
import random
from faker import Factory

wb = load_workbook('Excel水平测试题.xlsx')
sheet_data1 = wb['2-技巧中级']

list_row = list(range(6,20))
list_name = []
for i in list_row:
    list_name.append(sheet_data1['A' + str(i)].value)

list_sub = []
for i in range(len(list_row) - 1):
    if list_name[i] == list_name[i + 1]:
        list_sub.append(i + 6)
        print(list_sub)
    else:
        list_sub.append(i + 6)
        start_cell = 'A' + str(list_sub[0])
        end_cell = 'A' + str(list_sub[-1])
        interval_1 = start_cell + ':' + end_cell
        sheet_data1.merge_cells(interval_1)
        list_sub = []


print(list_name)

# for i in list_index:
#     list_col.append(get_column_letter(i))
#     list_col_write.append(get_column_letter(i + 10))
#
# dic_list = []
# dic_1 = {}
# set_list = []
# set_1 = set()
#
# for i in range(3,7):
#     for j , k in enumerate(list_col):
#         dic_1[j + 1] = sheet_data[k + str(i)].value
#         set_1.add(sheet_data[k + str(i)].value)
#     set_list.append(set_1)
#     dic_list.append(dic_1)
#     dic_1 = {}
#     set_1 = set()
#
# result_list = []
# result_list_1 = []
# for i in range(4):
#     list_pro = list(set_list[i])
#     list_pro.sort()
#     imp_num = list_pro[-2]
#
#     num = 0
#     for key, value in dic_list[i].items():
#         if value == imp_num:
#             sheet_data[list_col_write[num] + str(i + 3)].value = key
#             num += 1
#             result_list_1.append(key)
#     if result_list_1 != []:
#         result_list.append(result_list_1)
#     result_list_1 = []

# print(list_index)
# print(list_col)
# print(list_col_write)
# print(dic_list)
# print(result_list)
# print()
# print(set_list)
# print(list_data)

# index_num = []
# result_data = []
# data = 0
# count_1 = 0
# color_use = randomcolor()
# for i , indata in enumerate(list_data):
#     print(i)
#     data += indata
#     count_1 += 1
#     if (data > 11299 or count_1 > 7) and (i != 0):
#         data -= indata
#         result_data.append(data)
#         index_num.append('K' + str(i + 2))
#         sheet_data['K' + str(i + 2)].value = data
#         data = indata
#         count_1 = 0
#         fill = PatternFill("solid", fgColor=color_use)
#         sheet_data['K' + str(i + 2)].fill = fill
#         color_use = randomcolor()
#     elif i == 47:
#         result_data.append(data)
#         index_num.append('K' + str(count_1 + i + 1))
#         sheet_data['K' + str(count_1 + i + 1)].value = data
#         fill = PatternFill("solid", fgColor=color_use)
#         sheet_data['K' + str(count_1 + i + 1)].fill = fill
#
#     if i == 0:
#         fill = PatternFill("solid", fgColor=color_use)
#         sheet_data['G3'].fill = fill
#     else:
#         fill = PatternFill("solid", fgColor= color_use)
#         sheet_data['G' + str(i + 3)].fill = fill
#
#
# print(result_data)
# print(index_num)
# list_out = []
# for i in range(2,4):
#     count_zero = 0
#     for j in range(1, sheet_data.max_column + 1):
#         # print(get_column_letter(j))
#         if sheet_data[get_column_letter(j) + str(i)].fill.start_color.index == 4 and sheet_data[get_column_letter(j) + str(i)].value == 0 :
#             count_zero += 1
#     sheet_data[get_column_letter(2) + str(i)].value = count_zero
# print(list_out)
#
# result_str = sheet_data['C2'].value
# i = sheet_data['C2'].fill.start_color
# j = sheet_data['D2'].fill.start_color.index
# re.split('([省])', result_str)
# print(re.split('([省 | 市])', result_str))
# result_replace = result_str.replace('省', '省-').replace('市', '市-').replace('，', '').replace('。', '')
# result_split = result_replace.split('-')
# result_set = set()
# result_list = []
# result_str = ''
# for i in result_split:
#     if i not in result_set:
#         result_set.add(i)
#         result_list.append(i)
#         result_str += i
# print(re.split('[省 | 市]', result_str))
# print(sheet_G_price['A2'].value)
# print(sheet_G_price['A10'].value)
# print(sheet_B_data['B2'].value)
# list_checklist = list(range(2, 7))
# list_price = list(range(2, 40, 8))
#
# for num in range(len(list_checklist)):
#     i = list_checklist[num]
#     j = list_price[num]
#     print(sheet_Checklist['B' + str(i)].value)
#     sheet_G_price['A' + str(j)].value = '=清单!A' + str(i)
#     sheet_Checklist['B' + str(i)].hyperlink = "#组价!A" + str(j) #写入超链接
#     # link = "workbookEx.xlsx#sheet2!E5"
#     print(sheet_G_price['A' + str(j)].value)
# ws1 = wb.create_sheet("New")



# print(sheet_ranges[cai_shu_cell].value)
# index_cl = []
#
# for i in range(3,15):
#     index_cl.append(get_column_letter(i))
#
# print(index_cl)
#
# # print(index_cl[0])
# dis_num_list = []
# for j in range(len(index_cl)):
#     data_col = []
#     for i in range(2,22):
#         data_col.append(sheet_ranges[index_cl[j] + str(i)].value)
#         # data_col.reverse()
#     data_col.reverse()
#     # print(data_col)
#
#     dis_num = 0
#     for k in range(len(data_col)):
#         if data_col[k] == 0:
#             dis_num += 1
#         elif data_col[k] == None:  #对于单元格中为空值的情况，用None。
#             dis_num += 1
#         elif data_col[k] == 1:
#             break
#         else:
#             break
#     dis_num_list.append(dis_num)
#
# print(dis_num_list)
#
# for i in range(len(index_cl)):
#     sheet_ranges[index_cl[i] + str(34)].value = dis_num_list[i]
#
#
wb.save('sample.xlsx')
