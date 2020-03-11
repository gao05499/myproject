from openpyxl import load_workbook
import numpy as np
from collections import Counter
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter , column_index_from_string
import re
import random
from faker import Factory

wb = load_workbook('删除另一列出现的数据.xlsx',data_only=True)
sheet_data_all = wb['全部考生']
sheet_data_ignore = wb['不计成绩考生']
sheet_data_overage = wb['计算成绩的考生']

row_max = sheet_data_ignore.max_row
row_max_all = sheet_data_all.max_row
list_ignore = set()
list_overage = []

for i in range(2,row_max + 1):
    list_ignore.add(sheet_data_ignore['D' + str(i)].value)

for i in range(2,row_max_all + 1):
    if sheet_data_all['D' + str(i)].value in list_ignore:
        pass
    else:
        list_overage.append(sheet_data_all['D' + str(i)].value)

for i in range(2,len(list_overage) + 2):
    sheet_data_overage['E' + str(i)].value = list_overage[i - 2]

print(list_overage)
print(list_ignore)

wb.save('sample.xlsx')
