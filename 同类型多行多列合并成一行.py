from openpyxl import load_workbook
# import numpy as np
# from collections import Counter
# from openpyxl.styles.colors import Color
# from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter , column_index_from_string
# import re
# import random
# from faker import Factory

wb = load_workbook('同类型多行多列合并成一行.xlsx',data_only=True)
sheet_data = wb['Sheet1']
result_data = wb['Sheet2']

list_al = []
for i in range(3,7):
    list_al.append(get_column_letter(i))

row_max = 7
list_leader = []
list_list_sub = []
list_sub = []
new_row_num = 0
for i in range(1,row_max + 1):
    if i == 1:
        new_row_num += 1
        list_leader.append(sheet_data['B' + str(i)].value)
        for j in list_al:
            list_sub.append(sheet_data[j + str(i)].value)
    else:
        if sheet_data['B' + str(i)].value == sheet_data['B' + str(i - 1)].value:
            for j in list_al:
                list_sub.append(sheet_data[j + str(i)].value)
        else:
            list_list_sub.append(list_sub)
            list_sub = []
            new_row_num += 1
            list_leader.append(sheet_data['B' + str(i)].value)
            for j in list_al:
                list_sub.append(sheet_data[j + str(i)].value)

list_list_sub.append(list_sub)


for i in range(0,new_row_num):
    result_data['A' + str(i + 1)].value = list_leader[i]
    for j in range(2,len(list_list_sub[i]) + 2):
        col_num = get_column_letter(j)
        result_data[col_num + str(i + 1)].value = list_list_sub[i][j - 2]

print(len(list_list_sub))

wb.save('sample.xlsx')
